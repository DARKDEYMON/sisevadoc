from openpyxl import load_workbook
from openpyxl.utils.escape import unescape
from colorama import Fore, Back, Style, init
import pdb

class DNormalizador:
	def __init__(self, filename, data, receptor=None, number_worksheet=1, start_read_line=1, print_data=True, number_row_names=1):
		init(autoreset=True)
		self.wb = load_workbook(filename = filename)
		self.receptor = receptor
		self.number_worksheet = number_worksheet - 1
		self.start_read_line = start_read_line - 1
		self.print_data = print_data
		self.number_row_names = number_row_names
		self.normalizador(data)

	def worksheet(self):
		return self.wb.worksheets[self.number_worksheet]

	@staticmethod
	def formatoString(data):
		if(type(data)==str):
			return unescape(data).strip()
		return data

	@staticmethod
	def printcolor(number, string):
		colors = [Fore.GREEN, Fore.WHITE, Fore.CYAN, Fore.BLUE, Style.BRIGHT]
		print(colors[number-1] + "•" + str(DNormalizador.formatoString(string)))

	#funcion de retorno de la columna por numero de la fila actual
	def colOfRow(self, i):
		return lambda j : self.formatoString(self.worksheet().cell(row=i, column=j).value)

	#funcion de retorno el nombre de la columna por numero de la fila actual
	def colOfRowName(self):
		return lambda j : self.formatoString(self.worksheet().cell(row=self.number_row_names, column=j).value)

	#nombre de la columna actual 
	def nameOfCol(self, i):
		return self.formatoString(self.worksheet().cell(row=self.number_row_names, column=i).value)

	def __normalizador(self, n_colums_array, n_colum_array, data_comparison=[]):
		worksheet = self.worksheet()
		start = self.start_read_line
		n_level = len(data_comparison)+1
		if n_colum_array>=len(n_colums_array):
			return
		datas = []
		for i in range(1, worksheet.max_row + 1):
			data = worksheet.cell(row=i, column=n_colums_array[n_colum_array]).value
			if i>start:
				if(len(data_comparison)!=0):
					continue_loop = False
					for com in data_comparison:
						if worksheet.cell(row=i, column=com[0]).value != com[1]:
							continue_loop = True
					if continue_loop:
						continue
			
			if(i>start and data not in datas):
				if self.print_data:
					self.printcolor(n_level, data)
				c_data = [n_colums_array[n_colum_array],data]
				data_comparison_s = []
				data_comparison_s = data_comparison.copy()
				data_comparison_s.append(c_data)
				datas.append(data)
				#aqui ejecutar la interface esta es un afuncion con los siguientes datos (funcion de devolucion de nivel, dato, nomnre de columna, devolver fila por numero de la columna actual, devolver fila por nombre de la columna actual)
				if self.receptor:
					self.receptor(n_level, self.formatoString(data), self.nameOfCol(i), self.colOfRow(i), self.colOfRowName())
				self.__normalizador(n_colums_array, n_colum_array+1, data_comparison_s)

	def normalizador(self, n):
		self.__normalizador(n,0)
